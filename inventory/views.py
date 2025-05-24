from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Sum, F, Count
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Scooter, Parts, Store, StockTransfer, ScooterMaintenanceHistory, Supplier, Purchase, PurchaseItem
from .forms import (ScooterForm, PartsForm, StoreForm, StockTransferForm, MaintenanceHistoryForm,
                   SupplierForm, PurchaseForm, PurchaseItemForm, PurchaseItemFormSet)
from utils.export_utils import export_to_excel
from datetime import datetime
from users.utils import filter_by_user_store

# Scooter views
@login_required
def scooter_list(request):
    # Get all scooters and apply store-based access control
    scooters_queryset = Scooter.objects.all().select_related('store')
    scooters_queryset = filter_by_user_store(scooters_queryset, request.user)
    
    # Filter by status if specified
    status_filter = request.GET.get('status')
    if status_filter and status_filter != 'all':
        scooters_queryset = scooters_queryset.filter(status=status_filter)
    
    # Filter by category if specified
    category_filter = request.GET.get('category')
    if category_filter and category_filter != 'all':
        scooters_queryset = scooters_queryset.filter(category=category_filter)
    
    # Search by License No or VIN/Serial
    search_query = request.GET.get('search', '')
    if search_query:
        scooters_queryset = scooters_queryset.filter(
            license_number__icontains=search_query
        ) | scooters_queryset.filter(
            vin__icontains=search_query
        )
    
    # Export to Excel if requested
    if 'export' in request.GET:
        from utils.export_utils import export_to_excel
        
        # Define columns for export: (field_name, display_name)
        columns = [
            ('vin', 'VIN/Serial Number'),
            ('make', 'Make'),
            ('model', 'Model'),
            ('category', 'Category'),
            ('year', 'Year'),
            ('color', 'Color'),
            ('status', 'Status'),
            ('mileage', 'Mileage'),
            ('store.name', 'Store'),
            ('purchase_date', 'Purchase Date'),
            ('purchase_price', 'Purchase Price (R)'),
            ('last_maintenance', 'Last Maintenance'),
            ('notes', 'Notes')
        ]
        
        return export_to_excel(
            data=scooters_queryset,
            columns=columns,
            filename='Scooter_Inventory',
            title='Scooter Inventory Report',
            sheet_name='Scooters'
        )
    
    # Get all possible statuses for filter dropdown
    scooter_statuses = Scooter.STATUS_CHOICES
    
    # No pagination - all results are returned
    scooters = scooters_queryset
    
    return render(request, 'inventory/scooter_list.html', {
        'scooters': scooters, 
        'statuses': scooter_statuses,
        'current_status': status_filter or 'all',
        'current_category': category_filter or 'all',
        'search_query': search_query
    })

@login_required
def scooter_create(request):
    if request.method == 'POST':
        form = ScooterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Scooter added successfully.')
            return redirect('inventory:scooter_list')
    else:
        form = ScooterForm()
    
    return render(request, 'inventory/scooter_form.html', {'form': form, 'title': 'Add New Scooter'})

@login_required
def scooter_update(request, pk):
    scooter = get_object_or_404(Scooter, pk=pk)
    
    if request.method == 'POST':
        form = ScooterForm(request.POST, instance=scooter)
        if form.is_valid():
            form.save()
            messages.success(request, 'Scooter updated successfully.')
            return redirect('inventory:scooter_list')
    else:
        form = ScooterForm(instance=scooter)
    
    return render(request, 'inventory/scooter_form.html', {
        'form': form, 
        'title': 'Update Scooter',
        'scooter': scooter
    })

@login_required
def scooter_detail(request, pk):
    scooter = get_object_or_404(Scooter, pk=pk)
    maintenance_history = ScooterMaintenanceHistory.objects.filter(scooter=scooter).order_by('-maintenance_date')
    
    context = {
        'scooter': scooter,
        'maintenance_history': maintenance_history
    }
    
    return render(request, 'inventory/scooter_detail.html', context)

# Parts views
@login_required
def parts_list(request):
    # Get sort parameter from query string, default to 'part_number'
    sort_by = request.GET.get('sort', 'part_number')
    
    # Get store filter parameter from query string, default to None (all stores)
    store_id = request.GET.get('store', None)
    
    # Validate sort parameter to prevent injection
    valid_sort_fields = ['part_number', 'name', 'category', 'current_stock', 'reorder_level', 'unit_price']
    
    # Check if it's a reverse sort (descending)
    if sort_by.startswith('-') and sort_by[1:] in valid_sort_fields:
        sort_field = sort_by
    elif sort_by in valid_sort_fields:
        sort_field = sort_by
    else:
        # Default to part_number if invalid sort field
        sort_field = 'part_number'
    
    # Get all parts with sorting applied
    parts_query = Parts.objects.all().select_related('store')
    
    # Apply store-based access control for non-admin users
    parts_query = filter_by_user_store(parts_query, request.user)
    
    # Apply store filter if provided
    if store_id and store_id.isdigit():
        parts_query = parts_query.filter(store_id=int(store_id))
    
    # Search by Part Number or Name
    search_query = request.GET.get('search', '')
    if search_query:
        parts_query = parts_query.filter(
            part_number__icontains=search_query
        ) | parts_query.filter(
            name__icontains=search_query
        )
    
    # Apply sorting
    parts_queryset = parts_query.order_by(sort_field)
    
    # Get all stores for the store filter dropdown
    stores = Store.objects.all()
    
    # Export to Excel if requested
    if 'export' in request.GET:
        from utils.export_utils import export_to_excel
        
        # Get store name for the report
        store_name = "All Stores"
        if store_id and store_id.isdigit():
            try:
                selected_store = Store.objects.get(id=int(store_id))
                store_name = selected_store.name
            except Store.DoesNotExist:
                pass
        
        # Prepare additional info for the report
        additional_info = {
            'Total Parts': parts_queryset.count(),
            'Search Query': search_query if search_query else 'None',
            'Sort Order': sort_field.replace('_', ' ').title()
        }
        
        # Define columns for export: (field_name, display_name)
        columns = [
            ('part_number', 'Part Number'),
            ('name', 'Part Name'),
            ('category', 'Category'),
            ('store.name', 'Store Location'),
            ('current_stock', 'Current Stock'),
            ('reorder_level', 'Reorder Level'),
            ('unit_price', 'Unit Price (R)'),
            ('location_in_store', 'Location in Store'),
            ('description', 'Description')
        ]
        
        return export_to_excel(
            data=parts_queryset,
            columns=columns,
            filename='Parts_Inventory_Report',
            title='Parts Inventory Report',
            sheet_name='Parts Inventory',
            store_name=store_name,
            additional_info=additional_info
        )
    
    # No pagination - return all results
    parts = parts_queryset
    
    # Pass the current sort field and store filter to the template context
    context = {
        'parts': parts,
        'current_sort': sort_field,
        'stores': stores,
        'selected_store_id': store_id if store_id and store_id.isdigit() else None,
        'search_query': search_query
    }
    
    return render(request, 'inventory/parts_list.html', context)

@login_required
def parts_create(request):
    if request.method == 'POST':
        form = PartsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Part added successfully.')
            return redirect('inventory:parts_list')
    else:
        form = PartsForm()
    
    return render(request, 'inventory/parts_form.html', {'form': form, 'title': 'Add New Part'})

@login_required
def parts_update(request, pk):
    part = get_object_or_404(Parts, pk=pk)
    
    if request.method == 'POST':
        form = PartsForm(request.POST, instance=part)
        if form.is_valid():
            form.save()
            messages.success(request, 'Part updated successfully.')
            return redirect('inventory:parts_list')
    else:
        form = PartsForm(instance=part)
    
    return render(request, 'inventory/parts_form.html', {
        'form': form, 
        'title': 'Update Part',
        'part': part
    })

# Store views
@login_required
def store_list(request):
    stores_queryset = Store.objects.all()
    
    # Pagination - 9 items per page
    paginator = Paginator(stores_queryset, 9)
    page = request.GET.get('page')
    
    try:
        stores = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        stores = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        stores = paginator.page(paginator.num_pages)
        
    return render(request, 'inventory/store_list.html', {'stores': stores})

@login_required
def store_create(request):
    if request.method == 'POST':
        form = StoreForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Store added successfully.')
            return redirect('inventory:store_list')
    else:
        form = StoreForm()
    
    return render(request, 'inventory/store_form.html', {'form': form, 'title': 'Add New Store'})

@login_required
def store_update(request, pk):
    store = get_object_or_404(Store, pk=pk)
    
    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            messages.success(request, 'Store updated successfully.')
            return redirect('inventory:store_list')
    else:
        form = StoreForm(instance=store)
    
    return render(request, 'inventory/store_form.html', {
        'form': form, 
        'title': 'Update Store',
        'store': store
    })

# Stock Transfer views
@login_required
def stock_transfer_list(request):
    transfers_queryset = StockTransfer.objects.all().select_related('source_store', 'destination_store', 'part')
    
    # Apply store-based access control for non-admin users
    transfers_queryset = filter_by_user_store(transfers_queryset, request.user)
    
    # Export to Excel if requested
    if 'export' in request.GET:
        from utils.export_utils import export_to_excel
        
        # Define columns for export: (field_name, display_name)
        columns = [
            ('transfer_number', 'Transfer Number'),
            ('source_store.name', 'Source Store'),
            ('destination_store.name', 'Destination Store'),
            ('part.part_number', 'Part Number'),
            ('part.name', 'Part Name'),
            ('quantity', 'Quantity'),
            ('transfer_date', 'Transfer Date'),
            ('status', 'Status'),
            ('created_by.username', 'Created By'),
            ('notes', 'Notes')
        ]
        
        return export_to_excel(
            data=transfers_queryset,
            columns=columns,
            filename='Stock_Transfers',
            title='Stock Transfers Report',
            sheet_name='Transfers'
        )
    
    # Pagination - 9 items per page
    paginator = Paginator(transfers_queryset, 9)
    page = request.GET.get('page')
    
    try:
        transfers = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        transfers = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        transfers = paginator.page(paginator.num_pages)
    
    return render(request, 'inventory/stock_transfer_list.html', {'transfers': transfers})

@login_required
def stock_transfer_create(request):
    if request.method == 'POST':
        form = StockTransferForm(request.POST)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.created_by = request.user
            
            # Auto-generate transfer number (format: TRF-YYYYMMDD-XXXX)
            from datetime import datetime
            import random
            date_str = datetime.now().strftime('%Y%m%d')
            random_num = ''.join([str(random.randint(0, 9)) for _ in range(4)])
            transfer.transfer_number = f"TRF-{date_str}-{random_num}"
            
            # Check if source store has enough stock
            part = form.cleaned_data['part']
            source_store = form.cleaned_data['source_store']
            quantity = form.cleaned_data['quantity']
            
            if part.store == source_store and part.current_stock >= quantity:
                # Update stock levels
                part.current_stock -= quantity
                part.save()
                
                transfer.save()
                messages.success(request, 'Stock transfer initiated successfully.')
                return redirect('inventory:stock_transfer_list')
            else:
                messages.error(request, 'Insufficient stock in source store.')
    else:
        form = StockTransferForm()
    
    return render(request, 'inventory/stock_transfer_form.html', {'form': form, 'title': 'Create Stock Transfer'})

@login_required
def stock_transfer_update(request, pk):
    transfer = get_object_or_404(StockTransfer, pk=pk)
    old_status = transfer.status
    
    if request.method == 'POST':
        form = StockTransferForm(request.POST, instance=transfer)
        if form.is_valid():
            new_transfer = form.save(commit=False)
            
            # If status changed to completed, update destination store stock
            if old_status != 'completed' and new_transfer.status == 'completed':
                part = new_transfer.part
                
                # First check if the part with the same part number exists in the destination store
                try:
                    # Try to find the part with the SAME part number at the destination store
                    dest_part = Parts.objects.get(
                        part_number=part.part_number,
                        store=new_transfer.destination_store
                    )
                    # If found, update the stock but maintain all other attributes
                    dest_part.current_stock += new_transfer.quantity
                    
                    # Also ensure all other attributes match the source part 
                    # (in case they've been updated at the source)
                    dest_part.name = part.name
                    dest_part.description = part.description
                    dest_part.reorder_level = part.reorder_level
                    dest_part.unit_price = part.unit_price
                    dest_part.category = part.category
                    dest_part.location_in_store = part.location_in_store
                    
                    dest_part.save()
                except Parts.DoesNotExist:
                    # If not found, create a new part with the SAME part number
                    dest_part = Parts.objects.create(
                        part_number=part.part_number,  # Use the exact same part number
                        name=part.name,
                        description=part.description,
                        store=new_transfer.destination_store,
                        current_stock=new_transfer.quantity,
                        reorder_level=part.reorder_level,
                        unit_price=part.unit_price,
                        category=part.category,
                        location_in_store=part.location_in_store
                    )
            
            new_transfer.save()
            messages.success(request, 'Stock transfer updated successfully.')
            return redirect('inventory:stock_transfer_list')
    else:
        form = StockTransferForm(instance=transfer)
    
    return render(request, 'inventory/stock_transfer_form.html', {
        'form': form, 
        'title': 'Update Stock Transfer',
        'transfer': transfer
    })

# Supplier views
@login_required
def supplier_list(request):
    suppliers_queryset = Supplier.objects.all()
    
    # Export to Excel if requested
    if 'export' in request.GET:
        from utils.export_utils import export_to_excel
        
        # Define columns for export: (field_name, display_name)
        columns = [
            ('name', 'Supplier Name'),
            ('contact_person', 'Contact Person'),
            ('phone', 'Phone'),
            ('email', 'Email'),
            ('website', 'Website'),
            ('address', 'Address'),
            ('account_number', 'Account Number'),
            ('payment_terms', 'Payment Terms'),
            ('is_active', 'Active'),
            ('notes', 'Notes')
        ]
        
        return export_to_excel(
            data=suppliers_queryset,
            columns=columns,
            filename='Suppliers',
            title='Suppliers List',
            sheet_name='Suppliers'
        )
    
    # Pagination - 9 items per page
    paginator = Paginator(suppliers_queryset, 9)
    page = request.GET.get('page')
    
    try:
        suppliers = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        suppliers = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        suppliers = paginator.page(paginator.num_pages)
    
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers})

@login_required
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier added successfully.')
            return redirect('inventory:supplier_list')
    else:
        form = SupplierForm()
    
    return render(request, 'inventory/supplier_form.html', {'form': form, 'title': 'Add New Supplier'})

@login_required
def supplier_update(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Supplier updated successfully.')
            return redirect('inventory:supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    
    return render(request, 'inventory/supplier_form.html', {
        'form': form, 
        'title': 'Update Supplier',
        'supplier': supplier
    })

@login_required
def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    purchases = Purchase.objects.filter(supplier=supplier).order_by('-invoice_date')
    
    context = {
        'supplier': supplier,
        'purchases': purchases
    }
    
    return render(request, 'inventory/supplier_detail.html', context)

# Purchase views
@login_required
def purchase_list(request):
    purchases_queryset = Purchase.objects.all().select_related('supplier')
    
    # Apply store-based access control for non-admin users
    # For purchases, we need to filter related PurchaseItems
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.store:
        # Get the user's assigned store
        user_store = request.user.profile.store
        
        # Find purchases that have at least one item linked to the user's store
        store_purchase_ids = PurchaseItem.objects.filter(store=user_store).values_list('purchase_id', flat=True)
        purchases_queryset = purchases_queryset.filter(id__in=store_purchase_ids)
    
    # Export to Excel if requested
    if 'export' in request.GET:
        from utils.export_utils import export_to_excel
        
        # Define columns for export: (field_name, display_name)
        columns = [
            ('invoice_number', 'Invoice Number'),
            ('supplier.name', 'Supplier'),
            ('invoice_date', 'Invoice Date'),
            ('due_date', 'Due Date'),
            ('status', 'Status'),
            ('total_amount', 'Total Amount (R)'),
            ('amount_paid', 'Amount Paid (R)'),
            ('notes', 'Notes')
        ]
        
        return export_to_excel(
            data=purchases_queryset,
            columns=columns,
            filename='Purchases',
            title='Purchase Invoices',
            sheet_name='Invoices'
        )
    
    # Pagination - 9 items per page
    paginator = Paginator(purchases_queryset, 9)
    page = request.GET.get('page')
    
    try:
        purchases = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        purchases = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        purchases = paginator.page(paginator.num_pages)
    
    return render(request, 'inventory/purchase_list.html', {'purchases': purchases})

@login_required
def purchase_create(request):
    formset = None  # Initialize formset to avoid UnboundLocalError
    
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        formset = PurchaseItemFormSet(request.POST)  # Initialize formset for POST
        
        if form.is_valid():
            purchase = form.save(commit=False)
            purchase.created_by = request.user
            purchase.save()
            
            # Process the formset with the saved purchase instance
            formset = PurchaseItemFormSet(request.POST, instance=purchase)
            if formset.is_valid():
                purchase_items = formset.save(commit=False)
                
                # Calculate total amount and update inventory levels for each purchased item
                total_amount = 0
                for item in purchase_items:
                    # If store is not set, use the purchase default store
                    if not item.store and purchase.store:
                        item.store = purchase.store
                    
                    # Calculate item total
                    item_total = item.quantity * item.unit_price
                    total_amount += item_total
                    
                    if item.part and item.store:
                        # Update current stock of the part
                        item.part.current_stock += item.quantity
                        item.part.save()
                    
                    # Save the purchase item
                    item.save()
                
                # Update the purchase total amount
                purchase.total_amount = total_amount
                purchase.save()
                
                # Save any deleted items from the formset
                formset.save()
                
                messages.success(request, 'Purchase invoice added successfully and inventory levels updated.')
                return redirect('inventory:purchase_list')
            else:
                # If formset is invalid, delete the purchase object and show errors
                purchase.delete()
                for i, error_dict in enumerate(formset.errors):
                    if error_dict:
                        for field, errors in error_dict.items():
                            for error in errors:
                                messages.error(request, f"Item {i+1} - {field}: {error}")
                if formset.non_form_errors():
                    for error in formset.non_form_errors():
                        messages.error(request, f"Form Error: {error}")
        else:
            # Show specific form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            messages.error(request, 'Please correct the errors above.')
    else:
        form = PurchaseForm()
        formset = PurchaseItemFormSet()
    
    return render(request, 'inventory/purchase_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Create Purchase Invoice'
    })

@login_required
def purchase_update(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    
    if request.method == 'POST':
        form = PurchaseForm(request.POST, instance=purchase)
        if form.is_valid():
            form.save()
            
            # Process the formset
            formset = PurchaseItemFormSet(request.POST, instance=purchase)
            if formset.is_valid():
                # Track items that are removed to adjust inventory
                original_items = {item.id: item for item in purchase.items.all()}
                
                # Save the updated formset items
                purchase_items = formset.save(commit=False)
                
                # Update inventory for existing items that changed quantities
                for item in purchase_items:
                    if item.id and item.id in original_items:
                        if item.part and item.store:
                            # Adjust inventory based on quantity difference
                            quantity_diff = item.quantity - original_items[item.id].quantity
                            if quantity_diff != 0:
                                item.part.current_stock += quantity_diff
                                item.part.save()
                    elif item.part and item.store:  # New item added
                        # Add new item's quantity to inventory
                        item.part.current_stock += item.quantity
                        item.part.save()
                    
                    # Save the purchase item
                    item.save()
                
                # Handle deleted items - reduce inventory
                for form in formset.deleted_forms:
                    item_id = form.instance.id
                    if item_id in original_items:
                        item = original_items[item_id]
                        if item.part and item.store:
                            # Remove deleted item quantity from inventory
                            item.part.current_stock -= item.quantity
                            if item.part.current_stock < 0:
                                item.part.current_stock = 0
                            item.part.save()
                
                # Save formset to handle deletions
                formset.save()
                
                messages.success(request, 'Purchase invoice updated successfully and inventory levels adjusted.')
                return redirect('inventory:purchase_list')
            else:
                for error in formset.errors:
                    messages.error(request, error)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PurchaseForm(instance=purchase)
        formset = PurchaseItemFormSet(instance=purchase)
    
    return render(request, 'inventory/purchase_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Update Purchase Invoice',
        'purchase': purchase
    })

# API views for AJAX calls
@login_required
def part_detail_api(request, pk):
    """API endpoint to get part details for AJAX requests"""
    try:
        part = get_object_or_404(Parts, pk=pk)
        data = {
            'id': part.id,
            'name': part.name,
            'part_number': part.part_number,
            'description': part.description,
            'unit_price': float(part.unit_price),
            'current_stock': float(part.current_stock),
            'category': part.category
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def store_parts_api(request, store_id):
    """API endpoint to get parts for a specific store"""
    try:
        parts = Parts.objects.filter(store_id=store_id)
        data = []
        for part in parts:
            data.append({
                'id': part.id,
                'name': part.name,
                'part_number': part.part_number,
                'current_stock': float(part.current_stock),
                'unit_price': float(part.unit_price)
            })
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def purchase_detail(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    items = purchase.items.all()
    
    context = {
        'purchase': purchase,
        'items': items
    }
    
    return render(request, 'inventory/purchase_detail.html', context)

# Delete views for each model
@login_required
def scooter_delete(request, pk):
    scooter = get_object_or_404(Scooter, pk=pk)
    
    if request.method == 'POST':
        scooter.delete()
        messages.success(request, 'Scooter deleted successfully.')
        return redirect('inventory:scooter_list')
    
    return render(request, 'inventory/confirm_delete.html', {
        'object': scooter,
        'object_name': f"{scooter.make} {scooter.model} ({scooter.vin})",
        'title': 'Delete Scooter',
        'cancel_url': 'inventory:scooter_list'
    })

@login_required
def parts_delete(request, pk):
    part = get_object_or_404(Parts, pk=pk)
    
    if request.method == 'POST':
        part.delete()
        messages.success(request, 'Part deleted successfully.')
        return redirect('inventory:parts_list')
    
    return render(request, 'inventory/confirm_delete.html', {
        'object': part,
        'object_name': f"{part.name} ({part.part_number})",
        'title': 'Delete Part',
        'cancel_url': 'inventory:parts_list'
    })

@login_required
def store_delete(request, pk):
    store = get_object_or_404(Store, pk=pk)
    
    if request.method == 'POST':
        store.delete()
        messages.success(request, 'Store deleted successfully.')
        return redirect('inventory:store_list')
    
    return render(request, 'inventory/confirm_delete.html', {
        'object': store,
        'object_name': store.name,
        'title': 'Delete Store',
        'cancel_url': 'inventory:store_list'
    })

@login_required
def stock_transfer_delete(request, pk):
    transfer = get_object_or_404(StockTransfer, pk=pk)
    
    if request.method == 'POST':
        transfer.delete()
        messages.success(request, 'Stock transfer deleted successfully.')
        return redirect('inventory:stock_transfer_list')
    
    return render(request, 'inventory/confirm_delete.html', {
        'object': transfer,
        'object_name': transfer.transfer_number,
        'title': 'Delete Stock Transfer',
        'cancel_url': 'inventory:stock_transfer_list'
    })

@login_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        supplier.delete()
        messages.success(request, 'Supplier deleted successfully.')
        return redirect('inventory:supplier_list')
    
    return render(request, 'inventory/confirm_delete.html', {
        'object': supplier,
        'object_name': supplier.name,
        'title': 'Delete Supplier',
        'cancel_url': 'inventory:supplier_list'
    })

@login_required
def purchase_delete(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    
    if request.method == 'POST':
        # Adjust inventory levels by removing purchased quantities
        for item in purchase.items.all():
            if item.part and item.store:
                # Decrease inventory by the purchased quantity
                item.part.current_stock -= item.quantity
                # Ensure we don't go below zero
                if item.part.current_stock < 0:
                    item.part.current_stock = 0
                item.part.save()
        
        purchase.delete()
        messages.success(request, 'Purchase invoice deleted successfully and inventory levels adjusted.')
        return redirect('inventory:purchase_list')
    
    return render(request, 'inventory/confirm_delete.html', {
        'object': purchase,
        'object_name': purchase.invoice_number,
        'title': 'Delete Purchase Invoice',
        'warning': 'This will remove all purchased items from inventory. Are you sure you want to proceed?',
        'cancel_url': 'inventory:purchase_list'
    })

# Purchase Quote Views
@login_required
def purchase_quote(request):
    """View to create a purchase quote for ordering parts"""
    # Get stores based on user's access rights
    if request.user.is_superuser or (hasattr(request.user, 'profile') and not request.user.profile.store):
        # Admin users can see all stores
        stores = Store.objects.filter(is_active=True).order_by('name')
    elif hasattr(request.user, 'profile') and request.user.profile.store:
        # Regular staff can only see their assigned store
        stores = Store.objects.filter(id=request.user.profile.store.id, is_active=True)
    else:
        # Fallback (shouldn't normally happen)
        stores = Store.objects.none()
    
    # Get parts based on user's store assignment
    all_parts = Parts.objects.all().order_by('category', 'name')
    parts = filter_by_user_store(all_parts, request.user)
    
    # Generate a unique quote number
    import datetime
    today = datetime.datetime.now()
    quote_number = f"{today.strftime('%Y%m%d')}-{Parts.objects.count()}"
    
    context = {
        'parts': parts,
        'stores': stores,
        'today': today,
        'quote_number': quote_number
    }
    
    return render(request, 'inventory/purchase_quote.html', context)

@login_required
def export_purchase_quote(request):
    """Export selected parts as a purchase quote Excel file"""
    import json
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    if request.method == 'POST':
        # Get form data from the POST request
        items_data = json.loads(request.POST.get('items', '[]'))
        store_id = request.POST.get('store_id', '')
        quote_date = request.POST.get('quote_date', datetime.now().strftime('%Y-%m-%d'))
        quote_ref = request.POST.get('quote_ref', f"PQ-{datetime.now().strftime('%Y%m%d')}")
        
        # Extract part IDs and quantities
        part_ids = [item['id'] for item in items_data]
        quantities = {item['id']: item['quantity'] for item in items_data}
        
        # Get selected parts
        selected_parts = Parts.objects.filter(id__in=part_ids)
        
        # Get store name if a store was selected
        store_name = "All Stores"
        if store_id:
            try:
                store = Store.objects.get(id=store_id)
                store_name = store.name
            except Store.DoesNotExist:
                pass
        
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Quote"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True)
        title_font = Font(name='Arial', size=14, bold=True)
        normal_font = Font(name='Arial', size=11)
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define header fill
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Write title and quote information
        ws['A1'] = 'PURCHASE QUOTE'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A3'] = f'Quote Reference:'
        ws['B3'] = quote_ref
        ws['A3'].font = header_font
        
        ws['A4'] = f'Date:'
        ws['B4'] = quote_date
        ws['A4'].font = header_font
        
        ws['A5'] = f'Store:'
        ws['B5'] = store_name
        ws['A5'].font = header_font
        
        # Write column headers (row 7)
        headers = ['Part No.', 'Item Name', 'Quantity', 'Category']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Write rows for each part
        row_num = 8
        for part in selected_parts:
            part_id = str(part.id)
            quantity = quantities.get(part_id, 1)
            
            # Part Number
            cell = ws.cell(row=row_num, column=1)
            cell.value = part.part_number
            cell.font = normal_font
            cell.border = thin_border
            
            # Item Name
            cell = ws.cell(row=row_num, column=2)
            cell.value = part.name
            cell.font = normal_font
            cell.border = thin_border
            
            # Quantity (user-specified)
            cell = ws.cell(row=row_num, column=3)
            cell.value = quantity
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Category
            cell = ws.cell(row=row_num, column=4)
            cell.value = part.category
            cell.font = normal_font
            cell.border = thin_border
            
            row_num += 1
        
        # Auto-adjust columns width for better readability
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            # Set a minimum column width, then adjust based on content
            ws.column_dimensions[col_letter].width = 15
            
        # Set specific column widths
        ws.column_dimensions['B'].width = 40  # Item Name column wider
            
        # Create response with Excel data
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{quote_ref}_Purchase_Quote.xlsx"'
        
        return response
    
    else:
        # If accessed directly via GET, redirect to purchase quote page
        messages.error(request, 'Please select items first to generate a quote.')
        return redirect('inventory:purchase_quote')
        
# API for scooter details with store information
@login_required
def scooter_details_api(request):
    """API endpoint to get scooter details including store info"""
    try:
        scooter_id = request.GET.get('scooter_id')
        if not scooter_id:
            return JsonResponse({
                'success': False,
                'error': 'No scooter_id provided'
            })
            
        scooter = get_object_or_404(Scooter, pk=scooter_id)
        data = {
            'success': True,
            'id': scooter.id,
            'make': scooter.make,
            'model': scooter.model,
            'vin': scooter.vin,
            'status': scooter.status,
            'store_id': scooter.store_id,
            'store_name': scooter.store.name if scooter.store else None
        }
    except Exception as e:
        data = {
            'success': False,
            'error': str(e)
        }
    
    return JsonResponse(data)

@login_required
def part_detail_api(request, pk):
    """API endpoint to get part details"""
    try:
        part = Parts.objects.get(pk=pk)
        data = {
            'unit_price': float(part.unit_price),
            'current_stock': float(part.current_stock),
            'name': part.name,
            'part_number': part.part_number,
            'description': part.description
        }
        return JsonResponse(data)
    except Parts.DoesNotExist:
        return JsonResponse({'error': 'Part not found'}, status=404)

@login_required
def store_parts_api(request, store_id):
    """API endpoint to get parts filtered by store"""
    try:
        store = Store.objects.get(pk=store_id)
        parts = Parts.objects.filter(store=store).order_by('name')
        
        parts_data = []
        for part in parts:
            parts_data.append({
                'id': part.id,
                'name': part.name,
                'part_number': part.part_number,
                'description': part.description,
                'unit_price': float(part.unit_price),
                'current_stock': float(part.current_stock),
                'display_name': f"{part.part_number} - {part.name}"
            })
        
        return JsonResponse({'parts': parts_data})
    except Store.DoesNotExist:
        return JsonResponse({'error': 'Store not found'}, status=404)
