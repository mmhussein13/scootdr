"""Utility functions for exporting data to Excel with professional formatting"""
import datetime
import csv
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def export_to_excel(data, columns, filename, title=None, sheet_name=None, store_name=None, additional_info=None):
    """
    Export data to a professionally formatted Excel file
    
    Args:
        data: QuerySet or list of model objects containing the data to export
        columns: List of tuples (column_name, display_name)
        filename: Name of the file to export (without extension)
        title: Report title to display at the top
        sheet_name: Name of the Excel sheet
        store_name: Store name for the report header
        additional_info: Dictionary of additional information to display
        
    Returns:
        HttpResponse object with Excel file
    """
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or "Report"
    
    # Define professional styles
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Add title section
    if title:
        ws.merge_cells(f'A{current_row}:{get_column_letter(len(columns))}{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.border = border
        current_row += 2
        
        # Add metadata
        ws[f'A{current_row}'] = f"Export Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{current_row}'].font = data_font
        current_row += 1
        
        if store_name:
            ws[f'A{current_row}'] = f"Store: {store_name}"
            ws[f'A{current_row}'].font = data_font
            current_row += 1
            
        if additional_info:
            for key, value in additional_info.items():
                ws[f'A{current_row}'] = f"{key}: {value}"
                ws[f'A{current_row}'].font = data_font
                current_row += 1
                
        current_row += 1  # Empty row
    
    # Add header row
    for col_idx, (_, display_name) in enumerate(columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=display_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    current_row += 1
    
    # Add data rows
    for row_data in data:
        for col_idx, (col_name, _) in enumerate(columns, 1):
            # Get value from model object, handle nested attributes with dots
            if '.' in col_name:
                parts = col_name.split('.')
                value = row_data
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    else:
                        value = None
                        break
            else:
                # Fixed: Handle Django model objects properly
                if hasattr(row_data, col_name):
                    value = getattr(row_data, col_name)
                else:
                    value = ''
            
            # Format specific types
            if isinstance(value, datetime.datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, datetime.date):
                value = value.strftime('%Y-%m-%d')
            elif value is None:
                value = ''
                
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    return response